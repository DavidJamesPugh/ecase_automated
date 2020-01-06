r"""
Processing of data files downloaded from
eCaseDownloader into a format to have a pivot table access
All files are created and saved in ‘G:\eCase\Downloads’,
unless otherwise specified.
"""

import csv
import math
import os
import time
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook, Workbook

import constants
import styles


def care_level_list():
    """
    Takes reports downloaded with care_level_CSV(driver).
    Generates an excel file for
    each area with a list of residents and their care levels. 
    """

    care_level = rf'{constants.DOWNLOADS_DIR}\pod_Residents.csv'
    mcf_care_level = rf'{constants.DOWNLOADS_DIR}\pod_MCF.csv'

    try:
        checker = load_workbook(rf'{constants.OUTPUTS_DIR}\Care Levels\MCF - CareLevels.xlsx')
        checker.save(rf'{constants.OUTPUTS_DIR}\Care Levels\MCF - CareLevels.xlsx')
        checker.close()
    except FileNotFoundError:
        create_book = Workbook()
        create_book.save(rf'{constants.OUTPUTS_DIR}\Care Levels\MCF - CareLevels.xlsx')
        create_book.close()

    areas = ['HOUSE 1 - Hector', 'HOUSE 2 - Marion Ross',
             'HOUSE 3 - Bruce', 'HOUSE 4 - Douglas', 'HOUSE 5 - Henry Campbell',
             'Balmoral', 'Iona',
             'Hill', 'Glen Macky', 'Glen Taylor', 'Strathmore',
             'Glenmore', 'Terraces']

    mcf_care_level_pd = pd.read_csv(mcf_care_level)
    mcf_care_level_pd = mcf_care_level_pd.sort_values(
        by=['WingDescription', 'RoomDescription'], ascending=True)

    mcf_care_level_file = pd.ExcelWriter(rf'{constants.OUTPUTS_DIR}\Care Levels\MCF - CareLevels.xlsx')
    for area in areas[0:7]:
        mcf_care_level_pd[mcf_care_level_pd.WingDescription == area].to_excel(
            mcf_care_level_file, sheet_name=f'{area}', index=False)

    mcf_care_level_file.save()
    mcf_care_level_file.close()

    rv_care_level_pd = pd.read_csv(care_level)
    rv_care_level_pd = rv_care_level_pd.sort_values(
        by=['Block', 'Unit'], ascending=True)

    rv_care_level_file = pd.ExcelWriter(rf'{constants.OUTPUTS_DIR}\Care Levels\RLV - CareLevels.xlsx')
    for area in areas[7:13]:
        rv_care_level_pd[rv_care_level_pd.Block == area].to_excel(
            rv_care_level_file, sheet_name=f'{area}', index=False,
            columns=['Title', 'LastName', 'FirstName', 'Block', 'Unit'])

    rv_care_level_file.save()
    rv_care_level_file.close()

    care_level_file = pd.ExcelWriter(rf'{constants.OUTPUTS_DIR}\Care Levels\S3 Dementia - CareLevels.xlsx')
    dementia_pd = mcf_care_level_pd[mcf_care_level_pd.ResidentCareLevel == 'Stage 3 Dementia']
    dementia_pd[dementia_pd.WingDescription != 'HOUSE 5 - Henry Campbell'].to_excel(
        care_level_file, 'Dementia', index=False)
    care_level_file.save()
    care_level_file.close()

    dem_care_level_file = load_workbook(rf'{constants.OUTPUTS_DIR}\Care Levels\S3 Dementia - CareLevels.xlsx')
    mcf_care_level_file = load_workbook(rf'{constants.OUTPUTS_DIR}\Care Levels\MCF - CareLevels.xlsx')
    rlv_care_level_file = load_workbook(rf'{constants.OUTPUTS_DIR}\Care Levels\RLV - CareLevels.xlsx')

    dementia_sheet = dem_care_level_file['Dementia']
    styles.print_settings(dementia_sheet, [23, 23, 23, 24, 26], landscape=False)

    dem_care_level_file.save(rf'{constants.OUTPUTS_DIR}\Care Levels\S3 Dementia - CareLevels.xlsx')
    dem_care_level_file.close()

    for area in areas[0:7]:
        mcf_care_level_sheet = mcf_care_level_file[area]
        styles.print_settings(mcf_care_level_sheet, [23, 23, 25, 17, 18],
                              landscape=False)

    mcf_care_level_file.save(rf'{constants.OUTPUTS_DIR}\Care Levels\MCF - CareLevels.xlsx')
    mcf_care_level_file.close()

    for area in areas[7:13]:
        rlv_care_level_sheet = rlv_care_level_file[area]
        styles.print_settings(rlv_care_level_sheet, [5, 15, 24.5, 11, 4.9, 17.9],
                              one_page=False, landscape=False)

    rlv_care_level_file.save(rf'{constants.OUTPUTS_DIR}\Care Levels\RLV - CareLevels.xlsx')
    rlv_care_level_file.close()

    os.startfile(rf'{constants.OUTPUTS_DIR}\Care Levels\MCF - CareLevels.xlsx')
    os.startfile(rf'{constants.OUTPUTS_DIR}\Care Levels\RLV - CareLevels.xlsx')
    os.startfile(rf'{constants.OUTPUTS_DIR}\Care Levels\S3 Dementia - CareLevels.xlsx')

    os.remove(care_level)
    os.remove(mcf_care_level)


def ecase_data_import():
    r"""
    Takes reports downloaded with ecase_data(driver).
    Transfers all data into an excel file
    ‘J:\Quality Data\Clinical Data\eCaseData.xlsx’.
    Also formats this for use in pivot tables. 
    """
    count = 2
    date_format = '%Y-%m-%d'

    #  # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    #  # # File Titles & Dictionaries# # # 
    #  # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

    incident_types = {'Absconding': 'Absconding', 'PA unreas force': 'Challenging Behaviour',
                      'Physical Aggression': 'Challenging Behaviour',
                      'Verbal': 'Challenging Behaviour',
                      'Wandering': 'Wandering', 'Death': 'Death',
                      'Fall': 'Fall', 'Fall frac': 'Serious Harm',
                      'Fall unwit': 'Fall', 'Wit no inj': 'Fall',
                      'Eye, ear, nose, mouth & throat infections': 'Infection',
                      'Gastro Intestinal': 'Infection',
                      'Respiratory tract': 'Infection', 'Reproductive system': 'Infection',
                      'Skin and Wound': 'Infection', 'Urinary Tract Infection': 'Infection',
                      'Inj': 'Injury', 'Inj dis': 'Injury', 'SCRAPE': 'Injury',
                      'CUTLC': 'Injury', 'BURN': 'Injury', 'Unstageable': 'Pressure',
                      'STDI': 'Pressure', 'PUNCTURE': 'Injury', 'Skin Condition': 'Injury',
                      'Skin tear': 'Skin tear', 'SURGICAL': 'Injury', 'Venous': 'Injury',
                      'Arterial': 'Injury', 'Choking': 'Other',
                      'Other': 'Other', 'Pressure': 'Pressure',
                      'Stage 1': 'Pressure', 'Stage 2': 'Pressure',
                      'Skin Tear 1a': 'Skin Tear', 'Skin Tear 1b': 'Skin Tear',
                      'Skin Tear 21': 'Skin Tear', 'Skin Tear 2b': 'Skin Tear',
                      'Skin Tear 3': 'Skin Tear', 'Lesion': 'Skin Tear',
                      'Weight Loss': 'Weight Loss', 'Weight Gain': 'Weight Gain',
                      'No Weight Change': 'No Weight Change',
                      'WeightLoss': 'Weight Loss', ' Bruising': 'Bruise',
                      'Major Failure': 'Other'
                      }

    ecase_main_headers = ['Incident', 'Area', 'Resident Name', 'Date Of Incident',
                          'Incident Type', 'Type Info', 'Injury Info',
                          'Infection Intervention', 'On Admission',
                          'Organisms Identified', 'Weight Change',
                          'Number of Days', 'Severity', 'MedType']

    ecase_resident_headers = ['ID', 'Resident Name', 'Date Of Birth', 'Admission',
                              'Wing', 'Room', 'Marital Status', 'Age', 'Care Level',
                              'Nationality']

    ecase_admissions_headers = ['Resident Name', 'Admission Date']

    ecase_wound_headers = ['Incident', 'Area', 'Name', 'Date Of Incident',
                           'Incident Type', 'Location', 'Injury Info',
                           'Admitted With', 'WoundDays']

    house_list = ['ResidentInfo', 'Admissions',
                  'IncidentDatabase', 'WoundRecord']

    #  # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    #  # # File setup# # # 
    #  # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    ecase_main = load_workbook(rf'{constants.MAIN_DATA_DIR}\Clinical Data\eCaseData.xlsx')

    for house in house_list:
        if house in ecase_main.sheetnames:
            ecase_main.remove(ecase_main[house])
        ecase_main.create_sheet(house)

    ecase_wounds = ecase_main['WoundRecord']
    ecase_database = ecase_main['IncidentDatabase']
    ecase_admissions = ecase_main['Admissions']
    ecase_resident_info = ecase_main['ResidentInfo']

    alpha = []
    for letter in range(65, 91):
        alpha.append(chr(letter))

    for i in range(13):
        ecase_database[f'{alpha[i]}1'] = ecase_main_headers[i]

    for i in range(9):
        ecase_wounds[f'{alpha[i]}1'] = ecase_wound_headers[i]

    #  # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    #  # # Data Imports# # # 

    #  # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
    #  # # Weights Block# # # 
    #  # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
    weights_raw = []

    with open(rf'{constants.DOWNLOADS_DIR}\data_WeightChange.csv', newline='') as weights:
        weights_data = csv.reader(weights, delimiter=',', quotechar='"')
        for row in weights_data:
            weights_raw += [row]

    #  Weights entries into the eCase dBase
    for row in weights_raw[1:len(weights_raw)]:
        date_from = datetime.strptime(row[3], date_format)
        date_to = datetime.strptime(row[4], date_format)
        date_current = datetime.strptime(row[4], date_format)
        time_period = (((date_to - date_from).total_seconds()) / 60 / 60 / 24) + 1

        if time_period >= 6:
            if (float(row[5]) / time_period) < -0.06:
                weight_status = 'Weight Loss'

            elif (float(row[5]) / time_period) > 0.06:
                weight_status = 'Weight Gain'

            else:
                weight_status = 'No Weight Change'
        else:
            weight_status = 'No Weight Change'

        if weight_status == 'No Weight Change':
            continue

        if len(row[0]) > 11:
            if row[0] == 'HOUSE5-HenryCampbell':
                wing = (row[0][0:5] + ' ' + row[0][5] + ' ' + row[0][6] +
                        ' ' + row[0][7:12] + ' ' + row[0][12:20])

            elif row[0] == 'HOUSE2-MarionRoss':
                wing = (row[0][0:5] + ' ' + row[0][5] + ' ' + row[0][6] +
                        ' ' + row[0][7:13] + ' ' + row[0][13:17])
            else:
                wing = (row[0][0:5] + ' ' + row[0][5] + ' ' + row[0][6] + ' ' + row[0][7:len(row[0])])

        else:
            wing = row[0]

        try:
            incident_type = incident_types[weight_status]

        except KeyError:
            incident_type = 'NA'

        ecase_database[f'A{count}'] = incident_type
        ecase_database[f'B{count}'] = wing
        ecase_database[f'C{count}'] = row[1] + ' ' + row[2]
        ecase_database[f'D{count}'] = date_current
        ecase_database[f'E{count}'] = weight_status
        ecase_database[f'K{count}'] = row[5]
        ecase_database[f'L{count}'] = time_period
        ecase_database[f'M{count}'] = f'=K{count}/L{count}'

        count += 1

    #  # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
    #  # # Infection Block# # # 
    #  # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    infections_raw = []

    with open(rf'{constants.DOWNLOADS_DIR}\data_Infections_all.csv', newline='') as infections:
        infections_data = csv.reader(infections, delimiter=',', quotechar='"')
        for row in infections_data:
            infections_raw += [row]

    for row in infections_raw[1:len(infections_raw)]:
        date_current = datetime.strptime(row[3], '%Y-%m-%d %H:%M:%S.%f')

        try:
            incident_type = incident_types[row[4]]

        except KeyError:
            incident_type = 'NA'

        ecase_database[f'A{count}'] = incident_type
        ecase_database[f'B{count}'] = row[0]
        ecase_database[f'C{count}'] = row[1] + ' ' + row[2]
        ecase_database[f'D{count}'] = date_current

        for i in range(4, 10):
            ecase_database[f'{alpha[i]}{count}'] = row[i]

        count += 1

    #  # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    #  # # Incidents Block# # # 
    #  # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    incidents_raw = []

    with open(rf'{constants.DOWNLOADS_DIR}\data_MonthlyIncidents.csv', newline='') as incidents:
        incidents_data = csv.reader(incidents, delimiter=',', quotechar='"')
        for row in incidents_data:
            incidents_raw += [row]

    for row in incidents_raw[1:len(incidents_raw)]:
        date_current = datetime.strptime(row[3], date_format)

        try:
            incident_type = incident_types[row[4]]

        except KeyError:
            incident_type = 'NA'

        if row[4] in ['Inj dis', 'Inj', 'Other']:
            if row[6] == ' Bruising' or row[6] == 'Bruising':
                ecase_database[f'A{count}'] = incident_types[' Bruising']
            else:
                ecase_database[f'A{count}'] = incident_type
        else:
            ecase_database[f'A{count}'] = incident_type

        if row[4] == 'WeightLoss':
            continue

        ecase_database[f'B{count}'] = row[0]
        ecase_database[f'C{count}'] = row[1] + ' ' + row[2]
        ecase_database[f'D{count}'] = date_current

        for i in range(4, 7):
            ecase_database[f'{alpha[i]}{count}'] = row[i]

        count += 1

    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
    # # # Medication Data# # # 
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

    meds_raw = []

    with open(rf'{constants.DOWNLOADS_DIR}\data_Medication.csv', newline='') as medication:
        meds_data = csv.reader(medication, delimiter=',', quotechar='"')
        for row in meds_data:
            meds_raw += [row]

    for row in meds_raw[1:len(meds_raw)]:
        date_current = datetime.strptime(row[3], '%Y-%m-%d')

        incident_type = 'Medication'

        ecase_database[f'A{count}'] = incident_type
        ecase_database[f'B{count}'] = row[0]
        ecase_database[f'C{count}'] = row[1] + ' ' + row[2]
        ecase_database[f'D{count}'] = date_current
        ecase_database[f'E{count}'] = row[4]
        ecase_database[f'F{count}'] = row[5]

        count += 1

        # # # Data Imports# # # 
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    # # # Wounds Block# # # 
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

    wounds_raw = []
    wound_count = 2

    with open(rf'{constants.DOWNLOADS_DIR}\data_Wounds.csv', newline='') as wounds:
        wounds_data = csv.reader(wounds, delimiter=',', quotechar='"')
        for row in wounds_data:
            wounds_raw += [row]

    for row in wounds_raw[1:len(wounds_raw)]:
        date_current = datetime.strptime(row[3], '%Y-%m-%d %H:%M:%S.%f')

        try:
            incident_type = incident_types[row[4]]

        except KeyError:
            incident_type = 'NA'

        ecase_wounds[f'A{wound_count}'] = incident_type
        ecase_wounds[f'B{wound_count}'] = row[0]
        ecase_wounds[f'C{wound_count}'] = f'{row[1]} {row[2]}'
        ecase_wounds[f'D{wound_count}'] = date_current

        for i in range(4, 7):
            ecase_wounds[f'{alpha[i]}{wound_count}'] = row[i]
            if i == 6:
                ecase_wounds[f'{alpha[i + 1]}{wound_count}'] = None
                ecase_wounds[f'{alpha[i + 2]}{wound_count}'] = row[i + 1]

        wound_count += 1

    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
    # # # Resident Information# # # 
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
    residents_raw = []
    resident_count = 2

    with open(rf'{constants.DOWNLOADS_DIR}\data_Resident_Info.csv', newline='') as residents:
        residents_data = csv.reader(residents, delimiter=',', quotechar='"')
        for row in residents_data:
            residents_raw += [row]

    for i in range(10):
        ecase_resident_info[f'{alpha[i]}1'] = ecase_resident_headers[i]

    for row in residents_raw[1:len(residents_raw)]:
        date_of_birth = datetime.strptime(row[3], '%Y-%m-%d')

        ecase_resident_info[f'A{resident_count}'] = row[0]
        ecase_resident_info[f'B{resident_count}'] = f'{row[1]} {row[2]}'
        ecase_resident_info[f'C{resident_count}'] = date_of_birth
        ecase_resident_info[f'D{resident_count}'] = f'=VLOOKUP(B{resident_count},Admissions!A:B,2,FALSE)'
        ecase_resident_info[f'E{resident_count}'] = row[4]
        ecase_resident_info[f'F{resident_count}'] = row[5]
        ecase_resident_info[f'G{resident_count}'] = row[6]
        ecase_resident_info[f'H{resident_count}'] = f'=(NOW()-C{resident_count})/365'
        ecase_resident_info[f'I{resident_count}'] = row[7]
        ecase_resident_info[f'J{resident_count}'] = row[8]

        resident_count += 1

    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    # # # Admission Data# # # 
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    admissions_raw = []
    resident_count = 2

    with open(rf'{constants.DOWNLOADS_DIR}\data_Admission_Date.csv', newline='') as admissions:
        admissions_data = csv.reader(admissions, delimiter=',', quotechar='"')
        for row in admissions_data:
            admissions_raw += [row]

    ecase_admissions['A1'] = (ecase_admissions_headers[0])
    ecase_admissions['B1'] = (ecase_admissions_headers[1])

    for row in admissions_raw[2:len(admissions_raw)]:
        if row[6] == '' or row[6] == 'AccomEntryDate':
            continue

        admission_date = datetime.strptime(row[6], '%Y-%m-%d')

        ecase_admissions[f'A{resident_count}'] = f'{row[1]} {row[2]}'
        ecase_admissions[f'B{resident_count}'] = admission_date
        resident_count += 1

    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    # # # Closing files# # # 
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

    ecase_main.save(rf'{constants.MAIN_DATA_DIR}\Clinical Data\eCaseData.xlsx')
    ecase_main.close()

    os.remove(rf'{constants.DOWNLOADS_DIR}\data_Wounds.csv')
    os.remove(rf'{constants.DOWNLOADS_DIR}\data_MonthlyIncidents.csv')
    os.remove(rf'{constants.DOWNLOADS_DIR}\data_Infections_all.csv')
    os.remove(rf'{constants.DOWNLOADS_DIR}\data_WeightChange.csv')
    os.remove(rf'{constants.DOWNLOADS_DIR}\data_Resident_Info.csv')
    os.remove(rf'{constants.DOWNLOADS_DIR}\data_Admission_Date.csv')
    os.remove(rf'{constants.DOWNLOADS_DIR}\data_Medication.csv')


def bowel_import(wing: str):
    r"""
    Takes reports downloaded with bowel_Report(driver,wing),
    or prev_Bowel_Report (driver,wing).
    Transfers this data into an excel file in
    J:\Quality Data\Clinical Data\BowelsRecord.
    Formatted into a sheet for each area,
    as well as a sheet that has all areas in it. Pivot Table ready. 

    """
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
    # # # Adding bowels data to new file# # # 
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

    bowel_raw = []
    ecase_bowel_headers = ['Wing', 'Name', 'Date', 'Time', 'BowelType',
                           'Incontinent']
    alpha = []
    for letter in range(65, 91):
        alpha.append(chr(letter))

    time.sleep(1)

    ecase_bowels_record = load_workbook(
        rf'{constants.MAIN_DATA_DIR}\Clinical Data\eCaseBowelsRecord.xlsx')
    ecase_all = ecase_bowels_record['All']

    ecase_wing = ecase_bowels_record[wing]

    with open(rf'{constants.DOWNLOADS_DIR}\bowel_report.csv', newline='') as bowels:
        bowel_data = csv.reader(bowels, delimiter=',', quotechar='"')
        for row in bowel_data:
            bowel_raw += [row]

    for i in range(6):
        ecase_wing[f'{alpha[i]}1'] = ecase_bowel_headers[i]
        ecase_all[f'{alpha[i]}1'] = ecase_bowel_headers[i]

    for row in bowel_raw[1:len(bowel_raw)]:
        date_of_birth = datetime.strptime(row[3], '%Y-%m-%d')
        ecase_wing.append([row[0], f'{row[1]} {row[2]}',
                           date_of_birth, row[4], row[5], row[6]])
        ecase_all.append([row[0], f'{row[1]} {row[2]}',
                          date_of_birth, row[4], row[5], row[6]])

    ecase_bowels_record.save(rf'{constants.MAIN_DATA_DIR}\Clinical Data\eCaseBowelsRecord.xlsx')
    ecase_bowels_record.close()

    os.remove(rf'{constants.DOWNLOADS_DIR}\bowel_report.csv')


def bowel_setup():
    """
    Sets up an excel file for use in bowel_Import.
    Establishes a sheet for each area.
    """
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
    # # # File Setup for Bowel records# # # 
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

    try:
        ecase_bowels_record = load_workbook(
            rf'{constants.MAIN_DATA_DIR}\Clinical Data\eCaseBowelsRecord.xlsx')

    except FileNotFoundError:
        ecase_bowels_record = Workbook()

    house_list = ['All',
                  'HOUSE 1 - Hector',
                  'HOUSE 2 - Marion Ross',
                  'HOUSE 3 - Bruce',
                  'HOUSE 4 - Douglas',
                  'HOUSE 5 - Henry Campbell',
                  'Stirling', 'Iona',
                  'Balmoral', 'Braemar']

    for house in house_list:
        if house in ecase_bowels_record.sheetnames:
            ecase_bowels_record.remove(ecase_bowels_record[house])
        ecase_bowels_record.create_sheet(house)

    ecase_bowels_record.save(
        rf'{constants.MAIN_DATA_DIR}\Clinical Data\eCaseBowelsRecord.xlsx')
    ecase_bowels_record.close()


def bowel_report_cleanup():
    """
    Run after bowel_Import in order to save
    the working excel document with the current date.
    """
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
    # # # Creating a copy of the file with the date in name# # # 
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

    now = datetime.now()
    ecase_bowels_record = load_workbook(
        rf'{constants.MAIN_DATA_DIR}\Clinical Data\eCaseBowelsRecord.xlsx')
    ecase_bowels_record.save(
        rf'{constants.MAIN_DATA_DIR}\Clinical Data\BowelsRecord\eCaseBowelsRecord_{now.day}-{now.month}-{now.year}.xlsx')
    ecase_bowels_record.close()
    os.startfile(
        rf'{constants.MAIN_DATA_DIR}\Clinical Data\BowelsRecord\eCaseBowelsRecord_{now.day}-{now.month}-{now.year}.xlsx')


def care_plans_setup():
    """
    Sets up an excel file with a sheet per area.
    """

    try:
        careplans_file = load_workbook(rf'{constants.OUTPUTS_DIR}\Care Plans\eCaseCareplans.xlsx')

    except FileNotFoundError:
        careplans_file = Workbook()

    house_list = ['All', 'HOUSE 1 - Hector',
                  'HOUSE 2 - Marion Ross',
                  'HOUSE 3 - Bruce',
                  'HOUSE 4 - Douglas',
                  'HOUSE 5 - Henry Campbell',
                  'Stirling', 'Iona',
                  'Balmoral', 'Braemar']

    for house in house_list:
        if house in careplans_file.sheetnames:
            careplans_file.remove(careplans_file[house])
        careplans_file.create_sheet(house)

    careplans_file.save(rf'{constants.OUTPUTS_DIR}\Care Plans\eCaseCareplans.xlsx')
    careplans_file.close()


def careplan_import(wing):
    """
    Takes reports generated from ecase_Care_Plans_Audit(driver).
    Will use the file generated from care_Plans_Setup()
    and list all careplans for each area and resident.
    Places in each area’s sheet, as well as in the sheet for All areas
    """

    resident_count = 0
    time.sleep(1)

    careplans_file = load_workbook(rf'{constants.OUTPUTS_DIR}\Care Plans\eCaseCareplans.xlsx')
    ecase_wing = careplans_file[wing]
    ecase_all = careplans_file['All']

    try:
        with open(rf'{constants.DOWNLOADS_DIR}\cp_Care Plan Audit.csv', newline='') as careplans:
            careplans_data = csv.reader(careplans, delimiter=',', quotechar='"')
            for row in careplans_data:
                resident_count += 1
                ecase_wing.append([row[0], row[2], row[1], row[3],
                                   f'=IF(D{resident_count}+180=180,"",D{resident_count}+180)',
                                   row[4], row[5], row[6], row[7]])
                ecase_all.append([row[0], row[2], row[1], row[3],
                                  f'=IF(D{resident_count}+180=180,"",D{resident_count}+180)',
                                  row[4], row[5], row[6], row[7]])

    except FileNotFoundError:
        pass

    ecase_wing['E1'] = 'Next Due'
    ecase_all['E1'] = 'Next Due'

    careplans_file.save(rf'{constants.OUTPUTS_DIR}\Care Plans\eCaseCareplans.xlsx')

    xl = pd.ExcelFile(rf'{constants.OUTPUTS_DIR}\Care Plans\eCaseCareplans.xlsx')
    df = xl.parse(wing)
    df = df.sort_values(by=['Room', 'CarePlanName'])
    writer = pd.ExcelWriter(
        rf'{constants.MAIN_DATA_DIR}\Audits\CQR audits\{wing} Care Plans.xlsx')
    df.to_excel(writer, sheet_name=wing,
                columns=['Room', 'LastName', 'FirstName', 'DateCreated',
                         'Next Due', 'CarePlanName', 'CarePlanStatus', 'Author',
                         'Role'], index=False)
    writer.save()

    careplans_file.close()

    os.remove(rf'{constants.DOWNLOADS_DIR}\cp_Care Plan Audit.csv')


def careplans_missing_audits():
    r"""
    Opens the excel file made in care_Plans_Setup(), and careplan_Import(wing).
    Then saves a separate excel file for each area,
    with contents of only mandatory care plans and their status.
    For Lou, saves in ‘J:\Quality Data\Audits\CQR audits’. 
    """

    audits_required = ['About Me', 'Communication And Comprehension',
                       'Continence', 'Dietary', 'Leisure', 'Life History',
                       'Mobility and Transfers', 'PAS-CIS', 'Personal Hygiene',
                       'Pressure Injury Risk', 'Skin', 'Sleep']

    files = ['HOUSE 1 - Hector Care Plans.xlsx',
             'HOUSE 2 - Marion Ross Care Plans.xlsx',
             'HOUSE 3 - Bruce Care Plans.xlsx',
             'HOUSE 4 - Douglas Care Plans.xlsx',
             'HOUSE 5 - Henry Campbell Care Plans.xlsx',
             'Stirling Care Plans.xlsx', 'Iona Care Plans.xlsx',
             'Balmoral Care Plans.xlsx', 'Braemar Care Plans.xlsx']

    for file in files:
        audit_files = load_workbook(rf'{constants.MAIN_DATA_DIR}\Audits\CQR audits\{file}')

        if 'Missing' in audit_files.sheetnames:
            audit_files.remove(audit_files['Missing'])

        audit_files.create_sheet('Missing')
        audit_missing = audit_files['Missing']
        audit_missing.append(['Room Number', 'Audit Missing'])

        reader = pd.read_excel(rf'{constants.MAIN_DATA_DIR}\Audits\CQR audits\{file}',
                               index_col=0)
        for room in range(0, 31):
            try:
                # Index creates a list of all the unique index values in reader
                # Some rooms may be
                index = reader.index.unique()[room]
                careplans = reader.loc[[index], ['CarePlanName']]

                for audit in audits_required:
                    if audit not in careplans.values:
                        audit_missing.append([index, audit])
                audit_missing.append([])

            except IndexError:
                pass

        audit_files.save(rf'{constants.MAIN_DATA_DIR}\Audits\CQR audits\{file}')
        audit_files.close()


def doctor_numbers():
    """
    Takes the file downloaded with
    doctor_numbers_Download(driver).
    Establishes two new excel files,
    one with a summary of doctors allocation per area,
    and the other with a list of current residents
    with who their doctor is.  
    """

    count = 0
    doctor_header = ['Doctor', 'Wing']
    doctors_names = ['Mulgan', 'Hulley', 'Hodder', 'Mascher', 'JunJun']

    doc_file = open(rf'{constants.DOWNLOADS_DIR}\fixed_doctor_numbers.csv', 'w')
    doc_writer = csv.writer(doc_file)
    header = True

    with open(rf'{constants.DOWNLOADS_DIR}\doctor_numbers.csv', newline='') as doctors:
        doctors_data = csv.reader(doctors, delimiter=',', quotechar='"')
        for row in doctors_data:
            if header:
                doc_writer.writerow(doctor_header)
                header = False
                continue

            for name in doctors_names:
                if name in row[0]:
                    doc_writer.writerow([name, row[1]])

            if any(substring in row[0] for substring in doctors_names):
                continue
            else:
                doc_writer.writerow(['Other', row[1]])

    doc_file.close()

    doctors = pd.read_csv(rf'{constants.DOWNLOADS_DIR}\fixed_doctor_numbers.csv')
    doctors = doctors.pivot_table(index='Doctor', columns='Wing', aggfunc=len,
                                  margins=True).fillna(0).astype('int')
    doctors.to_excel(rf'{constants.OUTPUTS_DIR}\doctor_allocation_numbers.xlsx')

    doctors_sheet = load_workbook(rf'{constants.OUTPUTS_DIR}\doctor_allocation_numbers.xlsx')
    doctors_numbers = doctors_sheet['Sheet1']

    allocation = {'Mulgan': 37, 'Hulley': 65, 'Hodder': 38,
                  'Mascher': 40, 'JunJun': 40, 'Other': 0}

    # Add the allocation to the specific row of doctor in the excel file.
    # Every 11 cells, add the correct doctor's allocation according
    # to the docs name in the first column
    for i in doctors_numbers:
        for j in i:
            count += 1
            if count % 11 == 1:
                offset = math.ceil(count / 11)
                try:
                    doctors_numbers[f'L{offset}'] = [value for key, value in
                                                     allocation.items() if j.value in key][0]
                except IndexError:
                    pass

    doctors_numbers['L1'] = 'Allocation'

    widths = [13, 9, 9, 15, 20, 14, 16, 23, 5, 8, 10, 4, 12]
    styles.print_settings(doctors_numbers, widths)

    doctors_numbers.print_area = 'A1:M7'

    doctors_sheet.save(rf'{constants.OUTPUTS_DIR}\doctor_allocation_numbers.xlsx')

    # copyfile(rf'{constants.DOWNLOADS_DIR}\doctor_numbers.csv',
    #          rf'{constants.DOWNLOADS_DIR}\sorted_doctor_numbers.csv')

    doctors_sheet.close()
    os.remove(rf'{constants.DOWNLOADS_DIR}\doctor_numbers.csv')
    os.remove(rf'{constants.DOWNLOADS_DIR}\fixed_doctor_numbers.csv')
    # os.startfile(rf'\\{directory}\sorted_doctor_numbers.csv')
    os.startfile(rf'{constants.OUTPUTS_DIR}\doctor_allocation_numbers.xlsx')
