"""
-Processing of csv files downloaded from MYOB.
"""

import csv
import os

import pandas as pd
from openpyxl import Workbook, load_workbook

import constants
import styles


def training_lists():
    """
    Creates the mandatory training to be booked list.
    Places it in the directory specified in the function.
    Creates an individual sheet for each mandatory training,
    and puts the relevant employee information in each sheet.
    """

    myob_file = rf'{constants.STAFF_DIR}\training.CSV'
    training_file = pd.ExcelWriter(rf'{constants.MAIN_DATA_DIR}\Training\Training to be booked.xlsx')
        
    mandatory_training = ['Abuse and Neglect',
                          'Challenging Behaviours',
                          'Code of Rights',
                          'Continence Management',
                          'Cultural Safety',
                          'Death and Dying',
                          'Fire Safety',
                          'First Aid',
                          'Health and Safety-Disaster',
                          'Infection Control',
                          'Informed Consent',
                          'Man Handling & Falls Prevention',
                          'Manual Handling Non Clinical',
                          'Medication Management',
                          'Pain Management',
                          'Pressure Injuries',
                          'Privacy and Confidentiality',
                          'Restraint Minimisation',
                          'Sexuality and Intimacy',
                          'Skin Integrity and Grooming',
                          'Wound Care']

    training_data = pd.read_csv(myob_file, parse_dates=[4])
    # Removes the entries that have no booked date
    training_data = training_data[training_data.TLBOOKED != '/  /    ']
    # Converts the date to a datetime object so it sorts properly. NZ date format
    training_data.TLBOOKED = pd.to_datetime(training_data.TLBOOKED, format='%d/%m/%Y')
    # Sorts values by training and date
    sorted_data = training_data.sort_values(by=['TRNAME', 'TLBOOKED'], ascending=True)
    sorted_data['Name'] = sorted_data['FIRST_NAME'] + ' ' + sorted_data['LAST_NAME']
    sorted_data['TLBOOKED'] = pd.to_datetime(sorted_data['TLBOOKED'], format='%d%m%Y')
    # Accesses each sheet name in Training list above, and inserts the relevant entries
    for sheet in mandatory_training:
        if sheet == 'Man Handling & Falls Prevention':
            sorted_data[sorted_data.TRNAME == 'Manual Handling and Falls Prevention'].to_excel(training_file,
                                                                                               sheet, index=False)
        sorted_data[sorted_data.TRNAME == sheet].to_excel(training_file,
                                                          sheet,
                                                          index=False,
                                                          columns=['STAFF_CODE',
                                                                   'Name',
                                                                   'TRNAME',
                                                                   'TLBOOKED'])

    training_file.save()
    training_file.close()


def birthday_list():
    """
    Creates a birthday list, sorted by date,
    for this and the next two months.
    Will place in StaffDbase in Datatec folder
    """
    
    birthdays_raw = []
    birthdays_file = Workbook()
    birthdays_sheet = birthdays_file.active
    headers = ['Code', 'Name', 'Hours', 'Role', 'Start Date',
               'Birthdate', 'Area', 'Day', 'Month', 'Year']
    alpha = []

    for letter in range(65, 91):
        alpha.append(chr(letter))

    with open(rf'{constants.STAFF_DIR}\Birthday.CSV') as birthdays_info:
        birthdays_data = csv.reader(birthdays_info, delimiter=',', quotechar='"')
        for row in birthdays_data:
            try:
                month = int(str.split(row[5], sep='/')[1])
                day = int(str.split(row[5], sep='/')[0])
                year = int(str.split(row[5], sep='/')[2])
                birthdays_raw.append([row, day, month, year])

            except ValueError:
                month = 'Missing'
                day = 'Missing'
                year = 'Missing'
                birthdays_raw.append([row, day, month, year])

            except IndexError:
                pass
            
    for i in range(10):
        birthdays_sheet[f'{alpha[i]}1'] = headers[i]
                            
    for row in birthdays_raw:
        birthdays_sheet.append([row[0][0], row[0][1], row[0][2],
                                row[0][3], row[0][4],
                                row[0][5], row[0][6],
                                row[1], row[2], row[3]])

    birthdays_file.save(rf'{constants.STAFF_DIR}\StaffBirthdays.xlsx')
    birthdays_file.close()

    xl_file = pd.ExcelFile(rf'{constants.STAFF_DIR}\StaffBirthdays.xlsx')
    data_frame = xl_file.parse('Sheet')
    data_frame = data_frame.sort_values(by=['Month', 'Day'])

    writer = pd.ExcelWriter(rf'{constants.STAFF_DIR}\StaffBirthdays.xlsx')
    data_frame.to_excel(writer, sheet_name='Sheet',
                        columns=['Code', 'Name', 'Hours',
                                 'Role', 'Start Date',
                                 'Birthdate', 'Area',
                                 'Day', 'Month'], index=False)
    writer.save()
    writer.close()
    birthdays_file.save(rf'{constants.STAFF_DIR}\StaffBirthdays.xlsx')
    birthdays_file.close()
    file = load_workbook(rf'{constants.STAFF_DIR}\StaffBirthdays.xlsx')
    birthday_sheet = file.active
    widths = [5.6, 33.5, 29, 27, 11, 11, 20, 5, 7.5, 5]

    styles.print_settings(birthday_sheet, widths, 'Sheet')

    os.startfile(rf'{constants.STAFF_DIR}\StaffBirthdays.xlsx')
