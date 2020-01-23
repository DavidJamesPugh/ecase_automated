"""
-Processing and printing of admission/administrative files,
from eCase or from established files on the network
"""

import csv
import os
import re
from datetime import datetime
from zipfile import BadZipfile

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By

import button_functions
import constants
import styles


def pi_risk_levels(driver):
    """
    Takes the pir_code file with customer codes, and downloads and
    creates a file with all resident's PI risk factor
    """
    count = 1
    cust_codes = []
    codes_book = Workbook()
    codes_sheet = codes_book.active
    header = ['First Name', 'Last Name', 'Wing', 'Room', 'PI Risk', 'PI Desc']
    widths = [20, 15, 20, 7, 7, 15]
    styles.print_settings(codes_sheet, widths, header)

    with open(rf'{constants.DOWNLOADS_DIR}\pir_code.csv') as codes:
        codes_data = csv.reader(codes, delimiter=',')
        codes_data = list(codes_data)

        for row in codes_data[1:len(codes_data)]:
            try:
                if row[0] in cust_codes:
                    continue
                driver.get(
                    f'{constants.ECASE_URL}?action=assessment&accessmodule=assessment&customerCode={row[0]}')
                driver.implicitly_wait(10)
                driver.find_element_by_link_text('Expand All').click()
                driver.implicitly_wait(10)
                count += 1

            except NoSuchElementException:
                continue

            try:
                if driver.find_element_by_link_text('Pressure Injury Risk'):
                    driver.implicitly_wait(10)
                    pi = driver.find_element_by_link_text('Pressure Injury Risk')
                    driver.execute_script("arguments[0].click();", pi)
                    pi_risk = driver.find_element(By.XPATH, '//*[@id="PRRiskScore_span"]').text
                    pi_desc = driver.find_element(By.XPATH, '//*[@id="PRRiskScoreDesc_span"]').text
                    cust_codes += [row[0]]
                    codes_sheet[f'A{count}'] = row[1]
                    codes_sheet[f'B{count}'] = row[2]
                    codes_sheet[f'C{count}'] = row[3]
                    codes_sheet[f'D{count}'] = row[4]
                    codes_sheet[f'E{count}'] = pi_risk
                    codes_sheet[f'F{count}'] = pi_desc
                    codes_book.save(rf'{constants.OUTPUTS_DIR}\PIRiskLevels.xlsx')

            except NoSuchElementException:
                cust_codes += [row[0]]
                codes_sheet[f'A{count}'] = row[1]
                codes_sheet[f'B{count}'] = row[2]
                codes_sheet[f'C{count}'] = row[3]
                codes_sheet[f'D{count}'] = row[4]
                codes_sheet[f'E{count}'] = 'Not Done'
                codes_sheet[f'F{count}'] = 'Not Done'
                codes_book.save(rf'{constants.OUTPUTS_DIR}\PIRiskLevels.xlsx')

        codes_book.close()

    os.remove(rf'{constants.DOWNLOADS_DIR}\pir_code.csv')
    os.startfile(rf'{constants.OUTPUTS_DIR}\PIRiskLevels.xlsx')


def temp_movements_print():
    """
        Takes the csv report of Temporary movements from eCase,
        and appends any new movements that are related to Social Leave,
        Hospital, or Home visits.
        This is for Elizabeth Pane. 
    """

    movement_out = {}
    key_list = []
    headers = ['Key', 'First Name', 'Last Name', 'Wing', 'Room', 'Return date',
               'Days Away', 'Description', 'Leave Type']

    try:
        ecase_moves = load_workbook(rf'{constants.OUTPUTS_DIR}\eCaseTempMoves.xlsx')
        ecase_movements = ecase_moves['Temp Moves']

    except FileNotFoundError:
        ecase_moves = Workbook()
        ecase_movements = ecase_moves.active
        ecase_movements.title = 'Temp Moves'
        styles.print_settings(ecase_movements, header=headers)

    keys = ecase_movements['A':'A']
    for i in keys:
        key_list += [i.value]

    with open(rf'{constants.DOWNLOADS_DIR}\temp_movements.csv', newline='') as movements_info:
        movements_info_data = csv.reader(movements_info, delimiter=',', quotechar='"')
        movements_data = list(movements_info_data)
        for row in movements_data[1:len(movements_data)]:
            if (row[6] == 'Movement Out') and ('Death' not in row[7]):
                movement_out[row[0]] = [row[3], row[7]]
            else:
                #  This will check if a residentID is in the movement_Out dict,
                #  and will calculate how many days away they have been.
                if row[0] in movement_out:
                    if '-' in row[3]:
                        exit_date = str.split(movement_out[row[0]][0], '-')
                        entry_date = str.split(row[3], '-')
                        key_date = entry_date[0] + entry_date[1] + entry_date[2]
                        dash = True
                    else:
                        exit_date = str.split(movement_out[row[0]][0], '/')
                        entry_date = str.split(row[3], '/')
                        key_date = entry_date[0] + entry_date[1] + entry_date[2]
                        dash = False

                    if dash:
                        days_away = (datetime(int(entry_date[0]),
                                              int(entry_date[1]),
                                              int(entry_date[2])) -
                                     datetime(int(exit_date[0]),
                                              int(exit_date[1]),
                                              int(exit_date[2]))).days

                        entry_date = rf'{entry_date[2]}/{entry_date[1]}/{entry_date[0]}'

                    else:
                        days_away = (datetime(int(entry_date[2]),
                                              int(entry_date[1]),
                                              int(entry_date[0])) -
                                     datetime(int(exit_date[2]),
                                              int(exit_date[1]),
                                              int(exit_date[0]))).days

                        entry_date = rf'{entry_date[0]}/{entry_date[1]}/{entry_date[2]}'

                    data = [(row[0] + key_date), row[1], row[2], row[4], row[5],
                            entry_date, days_away, row[7], movement_out[row[0]][1]]

                    if row[0] + key_date not in key_list:
                        ecase_movements.append(data)

                    movement_out.pop(row[0])

    ecase_moves.save(rf'{constants.OUTPUTS_DIR}\eCaseTempMoves.xlsx')
    ecase_moves.close()
    os.startfile(rf'{constants.OUTPUTS_DIR}\eCaseTempMoves.xlsx')
    os.remove(rf'{constants.DOWNLOADS_DIR}\temp_movements.csv')


def create_front_sheet(village=False, no_print=False, nurses=False):
    """
        Takes the fs_Res and fs_Con reports from eCase,
        and produces a formatted front sheet for use in admission files.
        Prints out 2 copies with banking account information,
        and 1 without for the admission filing.
        This is for the Admission officer and accountants
    """

    basic_info_index = ['D6', 'D8', 'D9', 'D10', 'D12', 'D13', 'D14',
                        'D15', 'D16', 'I10', 'I13', 'I14',
                        'I15', 'I16']

    epoa_info_index = ['D21', 'D23', 'D24', 'D25', 'D26',
                       'I21', 'I23', 'I24', 'I25', 'I26']

    contact_info_index = ['D31', 'D33', 'D35', 'D36',
                          'D37', 'D40', 'D41', 'D42',
                          'D43', 'I31', 'I33', 'I35',
                          'I36', 'I37', 'I40', 'I41',
                          'I42', 'I43']

    funeral_info_index = ['D47', 'D48', 'I47',
                          'D51', 'D53', 'D54', 'D55',
                          'D57', 'D58', 'D59', 'D60',
                          'I51', 'I53', 'I54', 'I55',
                          'I57', 'I58', 'I59', 'I60']
    fresh_file = False
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

    try:
        sheet_book = load_workbook(rf'{constants.OUTPUTS_DIR}\front_sheet.xlsx')
        front_sheet = sheet_book.active

    # Catching an openpyxl exception, where i think the file wasnt saved
    # properly previously, and load_workbook causes this exception
    except BadZipfile:
        sheet_book = Workbook()
        front_sheet = sheet_book.active
        fresh_file = True

    except FileNotFoundError:
        sheet_book = Workbook()
        front_sheet = sheet_book.active
        fresh_file = True

    if fresh_file:
        #  headings
        main_heading_font = Font(size=14, bold=True, italic=True, color='000080')
        headings_font = Font(size=10, bold=True, italic=True, color='008000')
        sheet_titles_font = Font(size=10, bold=True, underline='single')

        sheet_headings = {'RESIDENTS INFORMATION FRONT SHEET': 'B4',
                          'ENDURING POWER OF ATTORNEY DETAILS': 'B19',
                          'CONTACTS FOR HEALTH AND WELFARE DECISIONS': 'B29',
                          'FUNERAL DIRECTOR': 'B46'}

        sheet_titles = {'Health and Welfare': 'B20', 'Property': 'G20',
                        'First Contact': 'B30', 'Second Contact': 'G30',
                        'Send Monthly SAV Account to': 'B50',
                        'Send Monthly Trust Account to': 'G50'}

        basic_info_fields = {'Location at SAV': 'B6', 'Title': 'B8',
                             'Surname': 'B9', 'Forenames': 'B10',
                             'Preferred Name': 'B11', 'Date of Birth': 'B12',
                             'Place of Birth': 'B13', 'Religion': 'B14',
                             'Gender': 'B15', 'Marital Status': 'B16',
                             'Doctor at SAV': 'G10', 'Telephone No.': 'G11',
                             'NHI No': 'G13', 'Date Admitted': 'G14',
                             'Care Level': 'G15', 'Ethnic Group': 'G16'}

        #  This is for two contacts
        epoa_info_fields = {'B21': 'Name', 'B23': 'Home Phone',
                            'B24': 'Work Phone', 'B25': 'Mobile Phone',
                            'B26': 'E-mail',
                            'G21': 'Name', 'G23': 'Home Phone',
                            'G24': 'Work Phone', 'G25': 'Mobile Phone',
                            'G26': 'E-mail'}

        #  This is for the Primary and secondary contacts
        contact_info_fields = {'B31': 'Name', 'B33': 'Relationship',
                               'B35': 'Address', 'B40': 'Home Phone',
                               'B41': 'Work Phone', 'B42': 'Mobile Phone',
                               'B43': 'E-mail',
                               'G31': 'Name', 'G33': 'Relationship',
                               'G35': 'Address', 'G40': 'Home Phone',
                               'G41': 'Work Phone', 'G42': 'Mobile Phone',
                               'G43': 'E-mail'}

        #  Funeral Director. Additional Monthly SAV and Trust account contact
        funeral_info_fields = {'B47': 'Company Name', 'B48': 'Phone Number',
                               'G47': 'Type of Service', 'B51': 'Name',
                               'B53': 'Address', 'B57': 'Home Phone',
                               'B58': 'Work Phone', 'B59': 'Mobile Phone',
                               'B60': 'E-mail',
                               'G51': 'Name', 'G53': 'Address',
                               'G57': 'Home Phone', 'G58': 'Work Phone',
                               'G59': 'Mobile Phone', 'G60': 'E-mail'}

        front_sheet['B1'].font = main_heading_font

        #  sheet headings writing to sheet & setting text styles
        for heading in sheet_headings:
            front_sheet[sheet_headings[heading]] = heading
            front_sheet[sheet_headings[heading]].font = headings_font

        #  sheet titles writing to sheet & setting text style
        for title in sheet_titles:
            front_sheet[sheet_titles[title]] = title
            front_sheet[sheet_titles[title]].font = sheet_titles_font

        # Writing the basic info headers into the file
        for info in basic_info_fields:
            front_sheet[basic_info_fields[info]] = info

        #  EPOA Details writing to sheet
        for epoa in epoa_info_fields:
            front_sheet[epoa] = epoa_info_fields[epoa]

        #  Contact info writing to sheet# # #
        for contact in contact_info_fields:
            front_sheet[contact] = contact_info_fields[contact]

        #  Funeral director info writing to sheet
        for funeral_info in funeral_info_fields:
            front_sheet[funeral_info] = funeral_info_fields[funeral_info]

        #  sheet image writing to sheet & positioning
        logo = Image(r'images\SAVLandscape.jpg')
        logo.anchor = 'A1'
        logo.width = 250
        logo.height = 40
        front_sheet.add_image(logo)
        sheet_book.save(rf'{constants.OUTPUTS_DIR}\front_sheet.xlsx')

        #  Setting text borders for whole sheet
        styles.full_border(front_sheet, 'D6')

        styles.full_border(front_sheet, 'D8:D11')
        styles.full_border(front_sheet, 'D12:D17')
        styles.full_border(front_sheet, 'I10:I11')
        styles.full_border(front_sheet, 'I13:I17')

        styles.full_border(front_sheet, 'D21')
        styles.full_border(front_sheet, 'D23:D26')
        styles.full_border(front_sheet, 'I21')
        styles.full_border(front_sheet, 'I23:I26')

        styles.full_border(front_sheet, 'D31')
        styles.full_border(front_sheet, 'D33')
        styles.full_border(front_sheet, 'D35:D38')
        styles.full_border(front_sheet, 'D40:D43')
        styles.full_border(front_sheet, 'I31')
        styles.full_border(front_sheet, 'I33')
        styles.full_border(front_sheet, 'I35:I38')
        styles.full_border(front_sheet, 'I40:I43')

        styles.full_border(front_sheet, 'D47:D48')
        styles.full_border(front_sheet, 'I47')

        styles.full_border(front_sheet, 'D51')
        styles.full_border(front_sheet, 'I51')
        styles.full_border(front_sheet, 'D53:D55')
        styles.full_border(front_sheet, 'I53:I55')
        styles.full_border(front_sheet, 'D57:D60')
        styles.full_border(front_sheet, 'I57:I60')

    #  Column widths
    styles.print_settings(front_sheet,
                          widths=[0.15, 17.0, .15,
                                  23.0, 4.15, 4.15,
                                  16.0, .15, 28.0],
                          landscape=False)

    respite = False

    #  Basic Resident info Writing to sheet# # # 
    doctors = ['Mascher', 'Jun', 'Mulgan', 'Hulley']

    if os.path.isfile(rf'{constants.DOWNLOADS_DIR}\p_name.txt'):
        p_file = open(rf'{constants.DOWNLOADS_DIR}\p_name.txt')
        p_name = p_file.read()
        p_file.close()
    else:
        p_name = ''

    front_sheet['D11'] = p_name

    with open(rf'{constants.DOWNLOADS_DIR}\fs_Res.csv', newline='') as basic_info:
        basic_info_data = csv.reader(basic_info, delimiter=',', quotechar='"')
        basic_data = list(basic_info_data)
        try:
            if 'Andrew' in basic_data[1][0]:
                basic_data[1][0] = basic_data[1][0][20:len(basic_data[1][0])]

            if not any(substring in basic_data[1][9] for substring in doctors):
                respite = True

            for cell in basic_info_index:
                front_sheet[cell] = basic_data[1][basic_info_index.index(cell)]
                if cell == 'D12':
                    front_sheet[cell] = (f'{basic_data[1][4][8:10]}/'
                                         f'{basic_data[1][4][5:7]}/'
                                         f'{basic_data[1][4][0:4]}')

                if cell == 'I14':
                    front_sheet[cell] = (f'{basic_data[1][11][8:10]}/'
                                         f'{basic_data[1][11][5:7]}/'
                                         f'{basic_data[1][11][0:4]}')

        except IndexError:
            if os.path.isfile(rf'{constants.DOWNLOADS_DIR}\fs_Res.csv'):
                os.remove(rf'{constants.DOWNLOADS_DIR}\fs_Res.csv')
            if os.path.isfile(rf'{constants.DOWNLOADS_DIR}\fs_Con.csv'):
                os.remove(rf'{constants.DOWNLOADS_DIR}\fs_Con.csv')
            return button_functions.popup_error("NHI is incorrect, please check you've entered it correctly "
                                                "and the resident is set up correctly with that NHI")

    for file in os.listdir(rf'{constants.DOWNLOADS_DIR}'):
        if re.match(r"^[A-Z]{3}[0-9]{4} Photo\.", file):
            photoname = file
            profile = Image(rf'{constants.DOWNLOADS_DIR}\{photoname}')
            profile.anchor = 'I2'
            profile.height = 140
            front_sheet.add_image(profile)
            sheet_book.save(rf'{constants.OUTPUTS_DIR}\front_sheet.xlsx')

    with open(rf'{constants.DOWNLOADS_DIR}\fs_Con.csv', newline='') as contact_info:
        contact_info_data = csv.reader(contact_info, delimiter=',', quotechar='"')
        contact_data = list(contact_info_data)
        for row in contact_data[1:len(contact_data)]:
            if row[9] == 'First Contact':
                for cell in contact_info_index[0:9]:
                    front_sheet[cell] = row[contact_info_index.index(cell)]

            elif row[9] == 'Second Contact':
                for cell in contact_info_index[9:18]:
                    front_sheet[cell] = row[contact_info_index.index(cell) - 9]

            elif row[9] == 'EPA Welfare':
                front_sheet[epoa_info_index[0]] = row[0]
                front_sheet[epoa_info_index[1]] = row[5]
                front_sheet[epoa_info_index[2]] = row[6]
                front_sheet[epoa_info_index[3]] = row[7]
                front_sheet[epoa_info_index[4]] = row[8]

            elif row[9] == 'EPA Property':
                front_sheet[epoa_info_index[5]] = row[0]
                front_sheet[epoa_info_index[6]] = row[5]
                front_sheet[epoa_info_index[7]] = row[6]
                front_sheet[epoa_info_index[8]] = row[7]
                front_sheet[epoa_info_index[9]] = row[8]

            elif row[9] == 'Funeral Director':
                front_sheet[funeral_info_index[0]] = row[0]
                front_sheet[funeral_info_index[1]] = row[6]

            elif row[9] == 'Send Fees Account' or row[9] == 'Billing':
                front_sheet[funeral_info_index[3]] = row[0]
                front_sheet[funeral_info_index[4]] = row[2]
                front_sheet[funeral_info_index[5]] = row[3]
                front_sheet[funeral_info_index[6]] = row[4]
                front_sheet[funeral_info_index[7]] = row[5]
                front_sheet[funeral_info_index[8]] = row[6]
                front_sheet[funeral_info_index[9]] = row[7]
                front_sheet[funeral_info_index[10]] = row[8]

            elif row[9] == 'Send Trust Account' or row[9] == 'Guaranator':
                front_sheet[funeral_info_index[11]] = row[0]
                front_sheet[funeral_info_index[12]] = row[2]
                front_sheet[funeral_info_index[13]] = row[3]
                front_sheet[funeral_info_index[14]] = row[4]
                front_sheet[funeral_info_index[15]] = row[5]
                front_sheet[funeral_info_index[16]] = row[6]
                front_sheet[funeral_info_index[17]] = row[7]
                front_sheet[funeral_info_index[18]] = row[8]

            elif row[9] == 'Resident':
                front_sheet['B17'] = 'Email'
                front_sheet['D17'] = row[8]
                front_sheet['G17'] = 'Contact Number'
                front_sheet['I17'] = row[5]

            # Doctors numbers.  SAV Drs dont want them on the front sheet anymore
            if respite:
                if row[9] == 'Medical Practitioner':
                    if row[7] != '':
                        front_sheet['I11'] = row[7]

                    elif row[6] != '':
                        front_sheet['I11'] = row[6]

                    elif row[5] != '':
                        front_sheet['I11'] = row[5]

                    else:
                        front_sheet['I11'] = 'No Number Present'
    try:
        #  Printing out Frontsheet without monthly accounts fields
        # For Everyone
        front_sheet.print_area = 'B1:I48'
        sheet_book.save(rf'{constants.OUTPUTS_DIR}\front_sheet.xlsx')
        if not no_print:
            os.startfile(rf'{constants.OUTPUTS_DIR}\front_sheet.xlsx', 'print')
            # Second info copy just for village nurses
            if nurses:
                os.startfile(rf'{constants.OUTPUTS_DIR}\front_sheet.xlsx', 'print')

        #  Printing out Frontsheet with monthly accounts fields
        # For Admissions and Village manager
        front_sheet.print_area = 'B1:I60'
        sheet_book.save(rf'{constants.OUTPUTS_DIR}\front_sheet.xlsx')
        if not no_print:
            if not nurses:
                os.startfile(rf'{constants.OUTPUTS_DIR}\front_sheet.xlsx', 'print')

        # Just for admissions
        if not no_print:
            if not village and not nurses:
                # print an extra accounts page if in the MCF
                os.startfile(rf'{constants.OUTPUTS_DIR}\front_sheet.xlsx', 'print')

        sheet_book.save(rf'{constants.OUTPUTS_DIR}\front_sheet.xlsx')
        sheet_book.close()

        info_files_remover()

    except PermissionError:
        info_files_remover()


def create_door_label(no_print=False):
    """
        Takes the fs_Res and fs_Con reports from eCase,
        and prints a formatted Door Label to place on the front
        of the resident’s room
    """

    try:
        try:
            sheet_book = load_workbook(rf'{constants.OUTPUTS_DIR}\door_label.xlsx')

        except BadZipfile:
            info_files_remover()
            return

        except FileNotFoundError:
            sheet_book = Workbook()
            sheet_book.save(rf'{constants.OUTPUTS_DIR}\door_label.xlsx')

        door_sheet = sheet_book.active

        if os.path.isfile(rf'{constants.DOWNLOADS_DIR}\p_name.txt'):
            p_file = open(rf'{constants.DOWNLOADS_DIR}\p_name.txt')
            p_name = p_file.read()
            p_file.close()
        else:
            p_name = ''

        with open(rf'{constants.DOWNLOADS_DIR}\fs_Res.csv', newline='') as basic_info:
            basic_info_data = csv.reader(basic_info, delimiter=',', quotechar='"')
            basic_data = list(basic_info_data)

        namecard_font = Font(size=36, bold=True, name='Arial')

        nhi_font = Font(size=28, bold=True, name='Copperplate Gothic Light')

        door_sheet['B6'] = basic_data[1][1] + ' ' + basic_data[1][3] + ' (' + p_name + ') ' + basic_data[1][2]
        door_sheet['B6'].font = namecard_font
        door_sheet['B6'].alignment = Alignment(horizontal='center', vertical='center',
                                               wrap_text=True)
        door_sheet.merge_cells(start_row=6, start_column=2, end_row=18, end_column=10)
        door_sheet['C24'] = basic_data[1][2]
        door_sheet['C24'].font = nhi_font
        door_sheet['C27'] = basic_data[1][1]
        door_sheet['C27'].font = nhi_font
        door_sheet['C29'] = basic_data[1][3]
        door_sheet['C29'].font = nhi_font
        door_sheet['C35'] = 'NHI No:'
        door_sheet['C35'].font = nhi_font
        door_sheet['F35'] = basic_data[1][10]
        door_sheet['F35'].font = nhi_font

        # # # Inserting Resident Photo
        for file in os.listdir(rf'{constants.DOWNLOADS_DIR}'):
            if re.match(r"^[A-Z]{3}[0-9]{4} Photo\.", file):
                photoname = file
                profile = Image(rf'{constants.DOWNLOADS_DIR}\{photoname}')
                profile.anchor = 'H21'
                profile.height = 212
                profile.width = 192
                door_sheet.add_image(profile)
                sheet_book.save(rf'{constants.OUTPUTS_DIR}\door_label.xlsx')

        styles.full_border(door_sheet, 'B6:J18', border=['double'])
        styles.full_border(door_sheet, 'B21:J38', border=['double'])

        door_sheet.print_area = 'A5:K39'
        styles.print_settings(door_sheet, landscape=False)
        sheet_book.save(rf'{constants.OUTPUTS_DIR}\door_label.xlsx')
        sheet_book.close()

        if not no_print:
            os.startfile(rf'{constants.OUTPUTS_DIR}\door_label.xlsx', 'print')

        info_files_remover()
    except PermissionError:
        info_files_remover()


def create_label_list():
    """
        Takes the fs_Res and fs_Con reports from eCase,
        and produces a formatted excel document for printing
        on a sheet of sticky labels. This won’t automatically
        print out the list, but it will open the formatted
        document to be printed from the bypass tray with the sticky labels.
        For Admissions officer and receptionist. 
    """
    try:
        try:
            sheet_book = load_workbook(rf'{constants.OUTPUTS_DIR}\label_sheet.xlsx')
        except FileNotFoundError:
            sheet_book = Workbook()
            sheet_book.save(rf'{constants.OUTPUTS_DIR}\label_sheet.xlsx')

        label_sheet = sheet_book.active
        doctors = ['Mascher', 'Jun', 'Mulgan', 'Hulley']

        styles.print_settings(label_sheet, widths=[14.714, 8.88571, 8.88571,
                                                   13.286, 11, 14.714, 8.88571,
                                                   8.88571, 13.286], landscape=False)

        respite = False

        if os.path.isfile(rf'{constants.DOWNLOADS_DIR}\p_name.txt'):
            p_file = open(rf'{constants.DOWNLOADS_DIR}\p_name.txt')
            p_name = p_file.read()
            p_file.close()
        else:
            p_name = ''

        with open(rf'{constants.DOWNLOADS_DIR}\fs_Res.csv', newline='') as basic_info:
            basic_info_data = csv.reader(basic_info, delimiter=',', quotechar='"')
            basic_data = list(basic_info_data)

            if not any(substring in basic_data[1][9] for substring in doctors):
                respite = True

        last_name = basic_data[1][2]

        if p_name == '':
            fore_names = basic_data[1][1] + ' ' + basic_data[1][3]
        else:
            fore_names = basic_data[1][1] + ' ' + basic_data[1][3] + f' ({p_name})'

        date_of_birth = (f'{basic_data[1][4][8:10]}/'
                         f'{basic_data[1][4][5:7]}/'
                         f'{basic_data[1][4][0:4]}')
        nhi = basic_data[1][10]
        gp = 'GP: ' + basic_data[1][9]
        sav = 'St Andrew\'s Village'
        room = basic_data[1][0]

        if respite:
            # #    Doctors numbers.  Drs dont want them on the labels anymore.
            # #         Except for respite.
            with open(rf'{constants.DOWNLOADS_DIR}\fs_Con.csv', newline='') as contact_info:
                contact_info_data = csv.reader(contact_info, delimiter=',', quotechar='"')
                contact_data = list(contact_info_data)
                for row in contact_data[1:len(contact_data)]:
                    if row[9] == 'Medical Practitioner':
                        gp = 'GP: ' + row[0]
                        if row[7] != '':
                            gp = gp + ' ' + row[7]

                        elif row[6] != '':
                            gp = gp + ' ' + row[6]

                        elif row[5] != '':
                            gp = gp + ' ' + row[5]

                        else:
                            gp = gp + ' ' + 'No Number Present'

        surname_font = Font(name='Arial', size=11, bold=True)
        forename_font = Font(name='Arial', size=11)
        med_norm_font = Font(name='Arial', size=10)
        small_bold_font = Font(name='Arial', size=10, bold=True)
        room_font = Font(name='Arial', size=7)

        left_list = ['', last_name, date_of_birth, gp, sav]
        right_list = ['', fore_names, nhi, '', room]

        for i in range(40):
            label_sheet.row_dimensions[i].height = float(21.25)

        for label in range(8):
            for label_row in range(1, 5):
                coeff = (label * 5) + label_row
                label_sheet[f'A{coeff}'] = left_list[label_row]
                label_sheet[f'F{coeff}'] = left_list[label_row]
                label_sheet[f'D{coeff}'] = right_list[label_row]
                label_sheet[f'I{coeff}'] = right_list[label_row]

                if label_row == 1:
                    label_sheet[f'A{coeff}'].font = surname_font
                    label_sheet[f'F{coeff}'].font = surname_font
                    label_sheet[f'D{coeff}'].font = forename_font
                    label_sheet[f'I{coeff}'].font = forename_font
                    label_sheet[f'D{coeff}'].alignment = Alignment(horizontal='right')
                    label_sheet[f'I{coeff}'].alignment = Alignment(horizontal='right')

                if label_row == 2:
                    label_sheet[f'A{coeff}'].font = med_norm_font
                    label_sheet[f'F{coeff}'].font = med_norm_font
                    label_sheet[f'D{coeff}'].font = small_bold_font
                    label_sheet[f'I{coeff}'].font = small_bold_font
                    label_sheet[f'D{coeff}'].alignment = Alignment(horizontal='right')
                    label_sheet[f'I{coeff}'].alignment = Alignment(horizontal='right')

                if label_row == 3:
                    label_sheet[f'A{coeff}'].font = med_norm_font
                    label_sheet[f'F{coeff}'].font = med_norm_font

                if label_row == 4:
                    label_sheet[f'A{coeff}'].font = small_bold_font
                    label_sheet[f'F{coeff}'].font = small_bold_font
                    label_sheet[f'D{coeff}'].font = room_font
                    label_sheet[f'I{coeff}'].font = room_font
                    label_sheet[f'D{coeff}'].alignment = Alignment(horizontal='right')
                    label_sheet[f'I{coeff}'].alignment = Alignment(horizontal='right')

        label_sheet.print_area = 'A1:I39'
        label_sheet.page_margins.top = .6
        label_sheet.page_margins.right = 0.27
        label_sheet.page_margins.bottom = .52
        label_sheet.page_margins.left = .48

        sheet_book.save(rf'{constants.OUTPUTS_DIR}\label_sheet.xlsx')
        sheet_book.close()

        os.startfile(rf'{constants.OUTPUTS_DIR}\label_sheet.xlsx')

        info_files_remover()

    except PermissionError:
        info_files_remover()


def village_birthdays(only_village=False):
    """
        Creates a formatted excel document of
        all Village residents with a birthday this month
    """

    birthdays_raw = []
    birthdays_file = Workbook()
    birthdays_sheet = birthdays_file.active
    current_date = datetime.now()
    widths = [5.6, 23, 23, 26, 12, 10, 10, 11.5, 4.5, 4.5, 8, 5]
    headers = ['Title', 'FirstName', 'LastName', 'Wing', 'Block', 'Unit',
               'Room', 'dateOfBirth', 'Age', 'Day', 'Month', 'Year']

    styles.print_settings(birthdays_sheet, widths=widths, header=headers)

    if current_date.month == 12:
        c_month = 1
        c_year = current_date.year + 1
    else:
        c_month = current_date.month + 1
        c_year = current_date.year

    with open(rf'{constants.DOWNLOADS_DIR}\birthdayList_MCF.csv') as birthdays_info:
        birthdays_data = csv.reader(birthdays_info, delimiter=',', quotechar='"')
        for row in birthdays_data:
            try:
                month = int(str.split(row[7], sep='-')[1])
                day = int(str.split(row[7], sep='-')[2])
                year = int(str.split(row[7], sep='-')[0])

                new_age = c_year - year
                if month >= c_month:
                    if month < (c_month + 1):
                        birthdays_raw.append([row, day, month, year, new_age])

                elif c_month == 12:
                    if month == 1 or month == 2:
                        birthdays_raw.append([row, day, month, year, new_age])

                elif c_month == 11:
                    if month == 1:
                        birthdays_raw.append([row, day, month, year, new_age])

            except IndexError:
                pass

    for row in birthdays_raw:
        birthdays_sheet.append([row[0][0], row[0][1], row[0][2], row[0][3], row[0][4],
                                row[0][5], row[0][6], row[0][7], row[4], row[1], row[2], row[3]])

    birthdays_file.save(rf'{constants.OUTPUTS_DIR}\Resident Birthdays\Residentbirthdays.xlsx')
    birthdays_file.close()

    xl = pd.ExcelFile(rf'{constants.OUTPUTS_DIR}\Resident Birthdays\Residentbirthdays.xlsx')
    df = xl.parse('Sheet')
    df = df.sort_values(by=['Month', 'Day'])
    writer = pd.ExcelWriter(rf'{constants.OUTPUTS_DIR}\Resident Birthdays\Residentbirthdays.xlsx')
    df.to_excel(writer, sheet_name='Sheet',
                columns=['Title', 'FirstName', 'LastName',
                         'Wing', 'Block', 'Unit', 'Room',
                         'dateOfBirth', 'Age',
                         'Day', 'Month', 'Year'], index=False)
    writer.save()
    writer.close()

    if only_village is False:
        os.startfile(rf'{constants.OUTPUTS_DIR}\Resident Birthdays\Residentbirthdays.xlsx')
        os.remove(rf'{constants.DOWNLOADS_DIR}\birthdayList_MCF.csv')

    # This continues the generation, and produces a file with only village
    # birthdays, formatted for use by Liz Keene
    else:
        xl = pd.ExcelFile(rf'{constants.OUTPUTS_DIR}\Resident Birthdays\Residentbirthdays.xlsx')
        df = xl.parse('Sheet')
        df = df.sort_values(by=['Block', 'Month', 'Day'])
        del df['Wing']
        del df['Room']
        df.dropna(axis=0, how='any', inplace=True)
        df = df[df.Block != 'Unknown']
        df = df[df.Block != '']

        writer = pd.ExcelWriter(rf'{constants.OUTPUTS_DIR}\Resident Birthdays\Villagerawbirthdays.xlsx')
        df.to_excel(writer, sheet_name='Sheet',
                    columns=['Title', 'FirstName', 'LastName',
                             'Block', 'Unit',
                             'dateOfBirth', 'Age', 'Day',
                             'Month', 'Year'], index=False)
        writer.save()
        writer.close()

        title = 'Resident birthdays'
        subtitle = 'Best Wishes from'
        subtitle2 = 'The Retirement Living Team!'
        date = datetime(c_year, c_month, 1).strftime("%B") + ' ' + str(c_year)

        birthdays_file = load_workbook(rf'{constants.OUTPUTS_DIR}\Resident Birthdays\Villagerawbirthdays.xlsx')
        birthday_sheet = birthdays_file.active

        village_birthdays_file = Workbook()
        resident_birthdays = village_birthdays_file.active

        title_font = Font(size=26, bold=True, name='Arial', color='703010')

        resident_birthdays['F2'] = title
        resident_birthdays['F3'] = date
        resident_birthdays['C46'] = subtitle
        resident_birthdays['C47'] = subtitle2

        resident_birthdays['F2'].font = title_font
        resident_birthdays['F2'].alignment = Alignment(horizontal='right',
                                                       vertical='center')
        resident_birthdays['F3'].font = title_font
        resident_birthdays['F3'].alignment = Alignment(horizontal='right',
                                                       vertical='center')
        resident_birthdays['C46'].alignment = Alignment(horizontal='center',
                                                        vertical='center')
        resident_birthdays['C47'].alignment = Alignment(horizontal='center',
                                                        vertical='center')

        # skip first 7 rows
        rowcount = 7
        alpha = []
        for letter in range(65, 91):
            alpha.append(chr(letter))

        for row in birthday_sheet.iter_rows(min_row=2):
            rowcount += 1
            count = 1
            month = datetime(row[9].value, row[8].value,
                             row[7].value).strftime("%B")

            data = [row[1].value + ' ' + row[2].value, '',
                    row[3].value + ' ' + row[4].value,
                    str(row[7].value) + ' ' + month, ' ', row[6].value]

            for element in data:
                resident_birthdays[f'{alpha[count]}{rowcount}'] = element
                count += 1

        styles.full_border(resident_birthdays, 'A2:F49', border=['thick'])
        styles.full_border(resident_birthdays, f'B8:E{rowcount}')
        resident_birthdays.print_area = "A2:F49"
        widths = [5.6, 27.5, 5.6, 16.9, 13, 10.5, 9.2]
        styles.print_settings(resident_birthdays, widths)
        village_birthdays_file.save(rf'{constants.OUTPUTS_DIR}\Resident Birthdays\VillageBirthdays-{date}.xlsx')
        village_birthdays_file.close()

        cakeimg = Image(rf'images\birthdaycake.jpg')
        cakeimg.height = 100
        cakeimg.width = 100
        messageimg = Image(rf'images\birthdayimage.jpg')
        messageimg.height = 100
        messageimg.width = 200

        village_birthdays_file = load_workbook(
            rf'{constants.OUTPUTS_DIR}\Resident Birthdays\VillageBirthdays-{date}.xlsx')
        resident_birthdays = village_birthdays_file.active

        resident_birthdays.add_image(messageimg, 'B2')
        village_birthdays_file.save(rf'{constants.OUTPUTS_DIR}\Resident Birthdays\Villagebirthdays-{date}.xlsx')
        resident_birthdays.add_image(cakeimg, 'C41')
        village_birthdays_file.save(rf'{constants.OUTPUTS_DIR}\Resident Birthdays\VillageBirthdays-{date}.xlsx')
        village_birthdays_file.close()

        os.startfile(rf'{constants.OUTPUTS_DIR}\Resident Birthdays\Villagebirthdays-{date}.xlsx')
        os.remove(rf'{constants.DOWNLOADS_DIR}\birthdayList_MCF.csv')


def info_files_remover():
    """

    :return:
    """
    for file in os.listdir(rf'{constants.DOWNLOADS_DIR}'):
        if re.match(r"^[A-Z]{3}[0-9]{4} Photo\.", file):
            os.remove(rf'{constants.DOWNLOADS_DIR}\{file}')
        if file in ['fs_Con.csv', 'fs_Res.csv', 'p_name.txt']:
            os.remove(rf'{constants.DOWNLOADS_DIR}\{file}')
