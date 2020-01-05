"""
-Data processing of the CSV files derived from the PDF of
Walls & Roche files from Clinical managers
"""

from datetime import datetime

from openpyxl import load_workbook, Workbook


def meds_counts():
    """
    Copy all the text from PDF and paste into a CSV file.
    Save and exit, and run this function on this new CSV.
    Creates an excel spreadsheet with all data sorted into a table,
    ready for a pivot table to be applied to it.
    """
    meds = load_workbook(r'J:\Quality Data\Data Technician\Walls and Roche\WRMedication.xlsx')

    new_meds = Workbook()
    new_meds_wr = new_meds['Sheet']
    new_meds_wr.title = 'Walls and Roche'

    row_count = 1
    alpha = []

    for letter in range(65, 91):
        alpha.append(chr(letter))

    for letter in range(65, 66):
        for letter2 in range(65, 91):
            a = chr(letter)
            b = chr(letter2)
            alpha.append(a + b)

    drugs = ['Aciclovir', 'Amoxicillin', 'AUGMENTIN', 'Cefaclor',
             'Cefalexin', 'CLINDAMYCIN', 'Clindamycin HCL',
             'Colecalciferol', 'Co-trimoxazole', 'DOXINE', 'Doxycycline',
             'Erythromycin', 'Flucloxacillin',
             'Fluconazole', 'Hexamine', 'HIPREX', 'Institution',
             'Methenamine hippurate', 'Metronidazole', 'Nitrofurantoin',
             'Norfloxacin', 'NORFLOXACIN', 'Ornidazole', 'Roxithromycin',
             'ROXITHROMYCIN', 'Trimethoprim', 'Valaciclovir', 'Ciprofloxacin',
             'Valganciclovir', 'Trimethoprim+Sulfamethoxazole']

    headers = ["Date", "id", "Doctor", "Resident", "NHI",
               "Area", "Qty", "Drug"]

    doctors = ['HULLEY', 'MASCHER', 'KidD', 'HODDER',
               'SHAW', 'POHL', 'MULGAN', 'CLEMO', 'LI',
               'CHONG', 'HEMMAWAY', 'KING', 'MACLACHLAN',
               'ANTUNOVICH', 'VANDERBOOR']
    date = 0

    column_count = 0
    area = 'St Andrews Village'
    for row in meds['Sheet1']:
        for cell in row:
            if cell.value is not None:
                try:
                    if type(cell.value) == datetime:
                        date = cell.value
                        continue

                    if 'Institution: St Andrews' in cell.value:
                        area = cell.value[24:len(cell.value)]
                        continue

                    if cell.value in ['Time \nMedicine', 'T/F ', 'Cd ', 'Rpt Prescriber ',
                                      'Cd Rpt Prescriber ', 'Patient', 'T ', 'X4 ', 'A4 ', '\n',
                                      'Prescription details report', 'St', 'Andrews',
                                      'RxNumber ', 'Qty ', 'Institution', 'group:']:
                        continue

                    if 'Take ' in cell.value:
                        continue

                    if 'Medicine' in cell.value or 'Time' in cell.value or 'Period: ' in cell.value:
                        continue

                    if '/' in cell.value and 'mg' not in cell.value and len(cell.value) < 40:
                        if len(cell.value) in [37, 38, 39]:
                            date = cell.value[15:23]
                            med_id = cell.value[23:33]
                            qty = cell.value[35:38]
                            column_count = 0
                            row_count += 1
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = date
                            column_count += 1
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = med_id
                            new_meds_wr[f'G{row_count}'] = qty
                            continue

                        elif len(cell.value) in [27, 28, 29]:
                            date = cell.value[5:13]
                            med_id = cell.value[14:24]
                            qty = cell.value[26:27]
                            column_count = 0
                            row_count += 1
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = date
                            column_count += 1
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = med_id
                            new_meds_wr[f'G{row_count}'] = qty
                            continue

                        elif len(cell.value) == 24:
                            date = cell.value[5:13]
                            med_id = cell.value[13:len(cell.value)]
                            column_count = 0
                            row_count += 1
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = date
                            column_count += 1
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = med_id
                            continue

                        elif len(cell.value) in [13, 14, 15]:
                            med_id = cell.value[0:10]
                            qty = cell.value[11:13]
                            row_count += 1
                            column_count = 0
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = date
                            column_count += 1
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = med_id
                            new_meds_wr[f'G{row_count}'] = qty
                            continue

                        elif len(cell.value) == 19:
                            date = cell.value[0:8]
                            med_id = cell.value[9:19]
                            row_count += 1
                            column_count = 0
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = date
                            column_count += 1
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = med_id
                            continue

                        elif len(cell.value) == 5:
                            date = cell.value
                            continue

                        row_count += 1
                        column_count = 0
                        new_meds_wr[f'{alpha[column_count]}{row_count}'] = date
                        column_count += 1
                        new_meds_wr[f'{alpha[column_count]}{row_count}'] = cell.value
                        continue

                    if len(cell.value) in [25, 26, 32, 33] and ':' in cell.value and 'NHI:' not in cell.value:
                        if 'Page' in cell.value or 'Institution' in cell.value or 'prescriptions' in cell.value:
                            continue
                        column_count = 0
                        row_count += 1
                        new_meds_wr[f'{alpha[column_count]}{row_count}'] = cell.value
                        row_count += 1

                        continue

                    for drug in drugs:
                        if drug in cell.value:
                            entry = cell.value.split()
                            if len(entry[0]) > 3:
                                column_count += 1

                            column_count += 1
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = entry[0]
                            column_count += 1
                            new_meds_wr[f'{alpha[column_count]}{row_count}'] = entry[1]
                            continue

                    for doctor in doctors:
                        if doctor in cell.value:
                            if column_count == 4:
                                column_count += 1
                                new_meds_wr[f'{alpha[column_count]}{row_count}'] = cell.value
                                doctor = new_meds_wr[f'F{row_count}'].value
                                temp = ['F', 'E', 'D', 'C']
                                for i in range(len(temp) - 1):
                                    new_meds_wr[f'{temp[i]}{row_count}'] = new_meds_wr[f'{temp[i + 1]}{row_count}'].value

                                new_meds_wr[f'C{row_count}'] = doctor

                            else:
                                column_count += 1
                                new_meds_wr[f'{alpha[column_count]}{row_count}'] = cell.value
                            continue

                    if 'NHI: ' in cell.value:
                        if len(cell.value) > 40:
                            doctorname = [doctor for doctor in doctors if doctor in cell.value]
                            med_id = cell.value[0:9]
                            name = cell.value[(len(doctorname) + 18):(len(cell.value) - 13)]
                            nhi = cell.value[(len(cell.value) - 7):len(cell.value)]
                            row_count += 1
                            column_count = 0
                            for i in [date, med_id, doctorname, name, nhi, area]:
                                try:
                                    new_meds_wr[f'{alpha[column_count]}{row_count}'] = i
                                    column_count += 1
                                except ValueError:
                                    pass

                        else:
                            name = cell.value[0:(len(cell.value) - 13)]
                            nhi = cell.value[(len(cell.value) - 7):len(cell.value)]
                            for i in [name, nhi, area]:
                                column_count += 1
                                new_meds_wr[f'{alpha[column_count]}{row_count}'] = i

                        continue

                except TypeError:
                    if type(cell.value) == int and cell.value > 2:
                        new_meds_wr[f'G{row_count}'] = cell.value
                        continue
    column_count = 0
    for elem in headers:
        new_meds_wr[f'{alpha[column_count]}1'] = elem
        column_count += 1

    date = datetime.today()
    date = (date.strftime("%d-%m-%y"))

    new_meds.save(rf'J:\Quality Data\Data Technician\Walls and Roche\new_meds - {date}.xlsx')
    new_meds.close()
