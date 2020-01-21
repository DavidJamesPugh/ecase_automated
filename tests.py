"""
    Unit Tests for the eCase migration package
"""
import csv
import os
import unittest

from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from selenium import webdriver

import constants
import downloader_support_functions
import ecase_downloader
import printing_documents
import styles


class TestingFrontSheet(unittest.TestCase):
    """
    Testing the front sheet docs
    """

    @classmethod
    def setUpClass(cls):
        """
        Downloading files and creating frontsheet
        :return:
        """
        nhi = 'PJY2787'

        cls.driver = ecase_downloader.ecase_login()
        ecase_downloader.resident_contacts(cls.driver, nhi)
        cls.driver.quit()

        with open(rf'{constants.DOWNLOADS_DIR}\fs_Res.csv') as data:
            cls.res_data = csv.reader(data, delimiter=',')
            cls.res_data = list(cls.res_data)[1]

        with open(rf'{constants.DOWNLOADS_DIR}\fs_Con.csv') as data:
            cls.con_data = csv.reader(data, delimiter=',')
            cls.con_data = list(cls.con_data)

        printing_documents.create_front_sheet(no_print=True)
        cls.front_sheet_book = load_workbook(rf'{constants.OUTPUTS_DIR}\front_sheet.xlsx')
        cls.front_sheet = cls.front_sheet_book['Sheet']

    def test_headers_present(self):
        sheet_headings = {'RESIDENTS INFORMATION FRONT SHEET': 'B4',
                          'ENDURING POWER OF ATTORNEY DETAILS': 'B19',
                          'CONTACTS FOR HEALTH AND WELFARE DECISIONS': 'B29',
                          'FUNERAL DIRECTOR': 'B46'}

        for header in sheet_headings:
            self.assertEqual(self.front_sheet[sheet_headings[header]].value,
                             header)

    def test_titles_present(self):
        sheet_titles = {'Health and Welfare': 'B20', 'Property': 'G20',
                        'First Contact': 'B30', 'Second Contact': 'G30',
                        'Send Monthly SAV Account to': 'B50',
                        'Send Monthly Trust Account to': 'G50'}

        for title in sheet_titles:
            self.assertEqual(self.front_sheet[sheet_titles[title]].value,
                             title)

    def test_basic_info_fields_present(self):

        basic_info_fields = {'Location at SAV': 'B6', 'Title': 'B8',
                             'Surname': 'B9', 'Forenames': 'B10',
                             'Preferred Name': 'B11', 'Date of Birth': 'B12',
                             'Place of Birth': 'B13', 'Religion': 'B14',
                             'Gender': 'B15', 'Marital Status': 'B16',
                             'Doctor at SAV': 'G10', 'Telephone No.': 'G11',
                             'NHI No': 'G13', 'Date Admitted': 'G14',
                             'Care Level': 'G15', 'Ethnic Group': 'G16'}

        for field in basic_info_fields:
            self.assertEqual(self.front_sheet[basic_info_fields[field]].value,
                             field)

    def test_epoa_info_fields_present(self):
        epoa_info_fields = {'B21': 'Name', 'B23': 'Home Phone',
                            'B24': 'Work Phone', 'B25': 'Mobile Phone',
                            'B26': 'E-mail',
                            'G21': 'Name', 'G23': 'Home Phone',
                            'G24': 'Work Phone', 'G25': 'Mobile Phone',
                            'G26': 'E-mail'}

        for field in epoa_info_fields:
            self.assertEqual(self.front_sheet[field].value,
                             epoa_info_fields[field])

    def test_contact_info_fields_present(self):
        contact_info_fields = {'B31': 'Name', 'B33': 'Relationship',
                               'B35': 'Address', 'B40': 'Home Phone',
                               'B41': 'Work Phone', 'B42': 'Mobile Phone',
                               'B43': 'E-mail',
                               'G31': 'Name', 'G33': 'Relationship',
                               'G35': 'Address', 'G40': 'Home Phone',
                               'G41': 'Work Phone', 'G42': 'Mobile Phone',
                               'G43': 'E-mail'}

        for field in contact_info_fields:
            self.assertEqual(self.front_sheet[field].value,
                             contact_info_fields[field])

    def test_funeral_info_fields_present(self):
        funeral_info_fields = {'B47': 'Company Name', 'B48': 'Phone Number',
                               'G47': 'Type of Service', 'B51': 'Name',
                               'B53': 'Address', 'B57': 'Home Phone',
                               'B58': 'Work Phone', 'B59': 'Mobile Phone',
                               'B60': 'E-mail',
                               'G51': 'Name', 'G53': 'Address',
                               'G57': 'Home Phone', 'G58': 'Work Phone',
                               'G59': 'Mobile Phone', 'G60': 'E-mail'}

        for field in funeral_info_fields:
            self.assertEqual(self.front_sheet[field].value,
                             funeral_info_fields[field])

    def test_res_data_present(self):
        basic_info_index = ['D6', 'D8', 'D9', 'D10', 'D12', 'D13', 'D14',
                            'D15', 'D16', 'I10', 'I13', 'I14',
                            'I15', 'I16']

        for index in range(len(basic_info_index)):
            if '-' in self.res_data[index] and len(self.res_data[index]) < 11:
                date = (f'{self.res_data[index][8:10]}/'
                        f'{self.res_data[index][5:7]}/'
                        f'{self.res_data[index][0:4]}')

                self.assertEqual(self.front_sheet[basic_info_index[index]].value,
                                 date)

            else:
                self.assertEqual(self.front_sheet[basic_info_index[index]].value,
                                 self.res_data[index])

    @classmethod
    def tearDownClass(cls):
        """

        :return:
        """
        pass


class TestingConstants(unittest.TestCase):
    """
    Testing whether the constant directories exist
    """

    def test_admission_dir(self):
        self.assertTrue(
            os.path.isdir(f'{constants.ADMISSION_DIR}'))

    def test_main_data_dir(self):
        self.assertTrue(
            os.path.isdir(f'{constants.MAIN_DATA_DIR}'))

    def test_downloads_dir(self):
        self.assertTrue(
            os.path.isdir(f'{constants.DOWNLOADS_DIR}'))

    def test_outputs_dir(self):
        self.assertTrue(
            os.path.isdir(f'{constants.OUTPUTS_DIR}'))

    def test_staff_dir(self):
        self.assertTrue(
            os.path.isdir(f'{constants.STAFF_DIR}'))

    def test_username(self):
        self.assertEqual(f'{constants.ECASE_USERNAME}', 'dpugh')

    def test_password(self):
        self.assertEqual(f'{constants.ECASE_PASSWORD}', 'sav2018')


class TestingEcase(unittest.TestCase):
    def setUp(self):
        """
        Connecting to google.com
        """
        self.driver = webdriver.Chrome()
        self.driver.get('https://sn.healthmetrics.co.nz/main.php?sx=0&')
        user_name = self.driver.find_element_by_id('mod_login_username')
        user_password = self.driver.find_element_by_id('mod_login_password')
        user_name.clear()
        user_name.send_keys(f'{constants.ECASE_USERNAME}')
        user_password.clear()
        user_password.send_keys(f'{constants.ECASE_PASSWORD}')
        self.driver.find_element_by_name('loginButton').click()


class TestingEcaseReportsAvailable(unittest.TestCase):
    """
    Tests that when you filter the reports in ecase, you only get the reports
    needed, and nothing extra
    """

    driver = None

    @classmethod
    def setUpClass(cls):
        """

        :return:
        """
        cls.driver = webdriver.Chrome()
        cls.driver.get(f'{constants.ECASE_URL}')
        user_name = cls.driver.find_element_by_id('mod_login_username')
        user_password = cls.driver.find_element_by_id('mod_login_password')
        user_name.clear()
        user_name.send_keys(f'{constants.ECASE_USERNAME}')
        user_password.clear()
        user_password.send_keys(f'{constants.ECASE_PASSWORD}')
        cls.driver.find_element_by_name('loginButton').click()

    def setUp(self):
        """
        Connecting to ecase and logging in. Navigate to report gen page
        """
        self.driver.get('https://sn.healthmetrics.co.nz/main.php?action=reportGenerator')

    def validate(self, prefix: str, reports_list: list):
        """
        Function to check if the reports in report_list exist in the
            ecase report generator under the prefix name filter
        :param prefix: string to filter report names
        :param reports_list: a list of which reports are meant to be in the
            report generator
        :return: Boolean on whether the report generator reports and report_list
            are the same
        """
        validation_list = []
        self.driver.find_element_by_id('filter-report-name').send_keys(f'{prefix}')
        self.driver.implicitly_wait(5)
        reports = self.driver.find_elements_by_css_selector('a.ng-binding')

        for line in reports:
            if line.text not in ['0', 'No', '']:
                validation_list += [line.text]

        reports_list.sort()
        validation_list.sort()
        self.assertEqual(reports_list, validation_list)

    def test_front_sheet_reports_downloadable(self):
        """
        if this fails, there is something wrong with the two reports needed
        for the front sheet info
        Either they have been deleted, or there are extra reports that will be
        downloaded unnecessarily
        :return:
        """
        self.validate('fs_', ['fs_Res', 'fs_Con'])

    def test_care_plans_reports_downloadable(self):
        """

        """
        self.validate('cp_', ['cp_Care Plan Audit'])

    def test_data_reports_downloadable(self):
        """

        """
        self.validate('data_', ['data_Admission_Date', 'data_Infections_all', 'data_Medication',
                                'data_MonthlyIncidents', 'data_Resident_Info', 'data_WeightChange',
                                'data_Wounds'])

    def test_doctor_number_reports_downloadable(self):
        """

        """
        self.validate('doctor_Numbers', ['doctor_Numbers'])

    def test_bowel_reports_downloadable(self):
        """

        """
        self.validate('bowel_report', ['bowel_report'])

    def test_pi_risk_reports_downloadable(self):
        """

        """
        self.validate('pir_code', ['pir_code'])

    def test_birthday_reports_downloadable(self):
        """

        """
        self.validate('birthdayList_MCF', ['birthdayList_MCF'])

    def test_podiatry_reports_downloadable(self):
        """

        """
        self.validate('pod_', ['pod_MCF', 'pod_Residents'])

    def test_temp_movements_reports_downloadable(self):
        """

        """
        self.validate('temp_movements', ['temp_movements'])

    @classmethod
    def tearDownClass(cls):
        """

        :return:
        """

        cls.driver.quit()


class PrintSettingsTests(unittest.TestCase):
    """

    """

    @classmethod
    def setUpClass(cls):
        """
        Instantiate a Testing_PrintSettings excel file to be used in the tests
        :return:
        """
        cls.test_book = Workbook()
        cls.test_book.save(r'unittests\Testing_PrintSettings.xlsx')
        cls.test_book.close()

    def setUp(self):
        """
        Setting up test workbook
        :return:
        """

        self.test_book = load_workbook(r'unittests\Testing_PrintSettings.xlsx')
        self.test_sheet = self.test_book.active
        logo = Image(r'images\SAVLandscape.jpg')
        logo.anchor = 'A1'
        logo.width = 250
        logo.height = 40
        self.test_sheet.add_image(logo)
        self.test_book.save(r'unittests\Testing_PrintSettings.xlsx')

    def test_column_widths_and_image(self):
        styles.print_settings(self.test_sheet,
                              [1, 2, 4, 8, 16, 32,
                               16, 21, 30, .5])

    def tearDown(self):
        """
        Adds a second image, and adds some text to help any troubleshooting
        :return:
        """
        logo = Image(r'images\SAVLandscape.jpg')
        logo.anchor = 'B10'
        logo.width = 250
        logo.height = 40
        self.test_sheet.add_image(logo)
        self.test_sheet['c5'] = 'There should be two SAV logos, one above this text and one below.'
        self.test_sheet['c6'] = 'The columns are set as [1, 2, 4, 8, 16, 32, 16, 21, 30, .5]'
        self.test_sheet['c7'] = 'The document should be landscape, and the printview should' \
                                'go to column J'
        self.test_sheet['j2'] = 'filler'
        self.test_book.save(r'unittests\Testing_PrintSettings.xlsx')
        self.test_book.close()


class StyleListParserTests(unittest.TestCase):
    """
    Testing the styles.list_parser
    """

    def test_no_value(self):
        self.assertRaises(IndexError, styles.list_parser, [])

    def test_single_value(self):
        test = styles.list_parser([1])
        self.assertEqual(test['bottom'], 1)
        self.assertEqual(test['top'], 1)
        self.assertEqual(test['left'], 1)
        self.assertEqual(test['right'], 1)

    def test_two_values(self):
        test = styles.list_parser(['hello', 'world'])
        self.assertEqual(test['bottom'], 'world')
        self.assertEqual(test['top'], 'world')
        self.assertEqual(test['left'], 'hello')
        self.assertEqual(test['right'], 'hello')

    def test_three_values(self):
        test = styles.list_parser(['goodbye', 'my', 'darling'])
        self.assertEqual(test['bottom'], 'my')
        self.assertEqual(test['top'], 'my')
        self.assertEqual(test['left'], 'darling')
        self.assertEqual(test['right'], 'goodbye')

    def test_four_values(self):
        test = styles.list_parser(['goodbye', 'my', 'danish', 'sweetheart'])
        self.assertEqual(test['bottom'], 'sweetheart')
        self.assertEqual(test['top'], 'my')
        self.assertEqual(test['left'], 'danish')
        self.assertEqual(test['right'], 'goodbye')

    def test_five_values(self):
        test = styles.list_parser([1, 2, 3, 4, 5])
        self.assertEqual(test['bottom'], 4)
        self.assertEqual(test['top'], 2)
        self.assertEqual(test['left'], 3)
        self.assertEqual(test['right'], 1)


class StyleBorderTests(unittest.TestCase):
    """
    Testing the styles.full_border
    """

    @classmethod
    def setUpClass(cls):
        """
        Instantiate a Testing_borders excel file to be used in the tests
        :return:
        """
        cls.test_book = Workbook()
        cls.test_book.save(r'unittests\Testing_Borders.xlsx')
        cls.test_book.close()

    def setUp(self):
        """
        Setting up test workbook
        :return:
        """

        self.test_book = load_workbook(r'unittests\Testing_Borders.xlsx')
        self.test_sheet = self.test_book.active
        self.test_book.save(r'unittests\Testing_Borders.xlsx')

    def test_one_row_full_border(self):
        """
        Testing a one row border
        :return:
        """
        styles.full_border(self.test_sheet, 'B2:G2')

    def test_one_cell_border(self):
        """
        Testing a one cell border
        :return:
        """
        styles.full_border(self.test_sheet, 'B4')

    def test_one_column_full_border(self):
        """
        Testing a one column border
        :return:
        """
        styles.full_border(self.test_sheet, 'B6:B12')

    def test_square_full_border(self):
        """
        Testing a square border
        :return:
        """
        styles.full_border(self.test_sheet, 'D4:G7')

    def test_rectangle_full_border(self):
        """
        Testing a rectangular border
        :return:
        """
        styles.full_border(self.test_sheet, 'D9:J10')

    def test_lower_case_cell_range(self):
        """
        Testing a cell range that was given in lowercase
        :return:
        """
        styles.full_border(self.test_sheet, 'i2:j7')

    def test_alternate_border(self):
        """
        Testing the border parameter by giving a 'double' linetype
        :return:
        """
        styles.full_border(self.test_sheet, 'D12:E18', border=['double'])

    def test_color_border(self):
        """
        Testing the rgb_hex parameter by passing in a RGB string
        :return:
        """
        styles.full_border(self.test_sheet, 'G12:J18', rgb_hex=['aaaaff'])

    def test_two_border_styles(self):
        styles.full_border(self.test_sheet, 'L2:M3', border=['double', 'thin'])

    def test_three_border_styles(self):
        styles.full_border(self.test_sheet, 'L5:M6', border=['double',
                                                             'thin',
                                                             'thick'])

    def test_four_border_styles(self):
        styles.full_border(self.test_sheet, 'L8:M9', border=['dashDot',
                                                             'dotted',
                                                             'hair',
                                                             'thick'])

    def test_four_border_styles_and_color(self):
        styles.full_border(self.test_sheet, 'L11:M12',
                           border=['dashDot', 'dotted', 'hair', 'thick', 'double'],
                           rgb_hex=['aa00ff', '00aaff', 'aaaa00', 'ffaa88'])

    def test_two_colors(self):
        styles.full_border(self.test_sheet, 'L14:M15',
                           rgb_hex=['aa00ff', '00aaff'])

    def test_three_colors(self):
        styles.full_border(self.test_sheet, 'O2:P3',
                           rgb_hex=['aa00ff', '00aaff', 'aaaa00'])

    def test_four_colors(self):
        styles.full_border(self.test_sheet, 'O5:P6',
                           rgb_hex=['aa00ff', '00aaff', 'aaaa00', 'ffaa88'])

    def test_no_values(self):
        styles.full_border(self.test_sheet, 'O8:P9',
                           border=[],
                           rgb_hex=[])

    def tearDown(self):
        r"""
        Printing what each box should be in each of the boxes corners
        Closing the workbook.
        :return: File unittests\Testing_Borders.xlsx with 8 boxes
        """
        self.test_sheet['B2'] = 'One Row'
        self.test_sheet['G2'] = 'One Row'

        self.test_sheet['B4'] = 'One Cell'

        self.test_sheet['B6'] = 'One Column'
        self.test_sheet['B12'] = 'One Column'

        self.test_sheet['D4'] = 'Square border'
        self.test_sheet['G7'] = 'Square border'

        self.test_sheet['I2'] = 'Rectangular column'
        self.test_sheet['J7'] = 'Rectangular column'

        self.test_sheet['D9'] = 'Rectangular Row'
        self.test_sheet['J10'] = 'Rectangular Row'

        self.test_sheet['D12'] = 'Double lined box'
        self.test_sheet['E18'] = 'Double lined box'

        self.test_sheet['G12'] = 'Coloured box "aaaaff"'
        self.test_sheet['J18'] = 'Coloured box "aaaaff"'

        self.test_sheet['L2'] = 'Two border types'
        self.test_sheet['L5'] = 'Three border types'
        self.test_sheet['L8'] = 'Four border types'

        self.test_sheet['L11'] = 'Four border types'
        self.test_sheet['L12'] = 'and color'

        self.test_sheet['L14'] = 'Two colors'
        self.test_sheet['O2'] = 'Three colors'
        self.test_sheet['O5'] = 'Four colors'

        self.test_sheet['O8'] = 'Empty values'

        self.test_book.save(r'unittests\Testing_Borders.xlsx')
        self.test_book.close()


class DateSelectorTests(unittest.TestCase):
    """
    Tests the date_selector function in eCaseDownloader.py
    """

    def test_past_month(self):
        self.assertEqual(downloader_support_functions.date_selector(1, 2019),
                         (6, 2, 4))

    def test_future_month(self):
        self.assertEqual(downloader_support_functions.date_selector(1, 2022),
                         (7, 6, 1))

    def test_current_datetime(self):
        self.assertEqual(downloader_support_functions.date_selector(12, 2019),
                         (7, 7, 2))


if __name__ == '__main__':
    unittest.main()
